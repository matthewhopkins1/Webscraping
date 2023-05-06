from urlib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from twilio.rest import Client
import keys

url = 'https://www.coingecko.com/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers = headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)

tables = soup.findAll('table')

crypto_table = tables[0]
crypto_rows = crypto_table.findAll('tr')

client = Client(keys.account_sid, keys.auth_token)
TwilioNum = ''
MyNum = ''

wb = xl.Workbook()
ws = wb.active
ws.title = 'Crytpo Prices'

ws['A1'] = 'Number'
ws['B1'] = 'Crypto Name and Symbol'
ws['C1'] = 'Current Price'
ws['D1'] = '24 Hour % Change'
ws['E1'] = '24 Hour Price Change'
ws['A1'].font = Font(size = 16, bold = True, color = '0000FF')
ws['B1'].font = Font(size = 20, bold = True, color = '0000FF')
ws['C1'].font = Font(size = 16, bold = False, color = '00FF00')
ws['D1'].font = Font(size = 16, bold = False, color = '0000ff')
ws['E1'].font = Font(size = 16, bold = False, color = '0000FF')

fill_cell = PatternFill(patternType= 'solid', fgColor= 'E2A3A3')
ws['A1'].fill = fill_cell
ws['B1'].fill = fill_cell
ws['C1'].fill = fill_cell
ws['D1'].fill = fill_cell
ws['E1'].fill = fill_cell

wb.save('PythontoExcel.xlsx')

for i in range(1,6):
    td = crypto_rows[i].findAll('td')
    number = td[1].text
    name = td[2].text
    price = float(td[3].text.replace(',', '').replace('$', ''))
    day_change = float(td[5].text.replace('%', ''))
    calc = round((price * 1 + day_change), 2)
    price_change = int(calc- price)
    if price_change <= 5 or price_change >= -5:
        text = client.messages.create(to = MyNum, from_ = TwilioNum, body = 'There has been a $5 change in price')
    print(text.status)

ws['A' + str(x+1)] = number
ws['B' + str(x+1)] = name
ws['C' + str(x+1)] = '$' + str(format(price, '.2f'))
ws['D' + str(x+1)] = str(format(day_change, ))
ws['E' + str(x+1)] = '$' + format(calc, '.2f')

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 22

wb.save('PythontoExcel.xlsx')

